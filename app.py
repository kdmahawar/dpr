import streamlit as st
import re
import os
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

# --- पेज सेटिंग ---
st.set_page_config(page_title="DPR Auto-Filler", layout="wide")
st.title("🚀 Quick DPR Generator")
st.markdown("##### Design & Concept : **K D Mahawar**")
st.markdown("---") 

TEMPLATE_FILE = "template.xlsx"
LAST_YEAR_FILE = "last_year_data.xlsx"

# --- ALIAS MAPPING ---
NAME_ALIASES = {
    "silica univ lts": "silica sand lts",
    "silica sand": "silica sand lts",
    "cumulative silica": "cumulative silica sand"
}

# --- HELPER FUNCTION: टेक्स्ट में से नंबर निकालना ---
def extract_float(text):
    if not text:
        return 0.0
    # अगर text में NIL लिखा है
    if "nil" in text.lower():
        return 0.0
    # नंबर ढूँढें (जिसमें डॉट भी हो सकता है)
    match = re.search(r"(\d+(\.\d+)?)", text)
    if match:
        return float(match.group(1))
    return 0.0

raw_text = st.text_area("WhatsApp Message यहाँ पेस्ट करें:", height=300)

if st.button("Excel फाइल बनाएँ"):
    if not os.path.exists(TEMPLATE_FILE):
        st.error(f"⚠️ Error: '{TEMPLATE_FILE}' नहीं मिली!")
    elif not raw_text:
        st.warning("⚠️ कृपया मैसेज पेस्ट करें।")
    else:
        try:
            wb = load_workbook(TEMPLATE_FILE)
            ws = wb.active
            
            # ---------------------------------------------------------
            # PART A: तारीख (Date)
            # ---------------------------------------------------------
            date_pattern = r"Date:.*?(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})"
            date_match = re.search(date_pattern, raw_text, re.IGNORECASE)
            
            final_date_str = "Unknown"
            lookup_date_obj = None
            
            if date_match:
                day, month, year = date_match.groups()
                if len(year) == 2: year = "20" + year
                
                final_date_str = f"{day.zfill(2)}-{month.zfill(2)}-{year}"
                lookup_date_obj = pd.to_datetime(f"{day}-{month}-{int(year)-1}", dayfirst=True)
                
                for row in ws.iter_rows(min_row=1, max_row=10):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and "Date:" in cell.value:
                            cell.value = f"Date: {final_date_str}"
                            break

            # ---------------------------------------------------------
            # PART B: पिछले साल का डेटा
            # ---------------------------------------------------------
            if lookup_date_obj and os.path.exists(LAST_YEAR_FILE):
                try:
                    ly_df = pd.read_excel(LAST_YEAR_FILE)
                    ly_df['Date'] = pd.to_datetime(ly_df['Date'], dayfirst=True)
                    target_row = ly_df[ly_df['Date'] == lookup_date_obj]
                    
                    if not target_row.empty:
                        ws['G6'] = target_row['Ball Clay'].values[0]
                        ws['G7'] = target_row['Silica'].values[0]
                        st.info(f"✅ Last Year Data ({lookup_date_obj.strftime('%d-%m-%Y')}) Found!")
                except Exception:
                    pass

            # ---------------------------------------------------------
            # PART C: व्हाट्सएप डेटा (ADVANCED REGEX)
            # ---------------------------------------------------------
            # अब हम strict number की जगह (.*?) का यूज़ कर रहे हैं, यानी "कुछ भी" उठा लो
            pattern = (
                r"(?:^|\n)\s*(?:\*)?([^\n\r*]+?)(?::)?(?:\*)?\s*\n\s*" 
                r"(?:•\s*)?Daily:\s*(.*?)\n\s*"     # कुछ भी टेक्स्ट कैप्चर करो (NIL, empty, numbers)
                r"(?:•\s*)?Monthly:\s*(.*?)\n\s*"   # Monthly का टेक्स्ट
                r"(?:•\s*)?Yearly:\s*(.*?)(?:\n|$)" # Yearly का टेक्स्ट
            )
            matches = re.findall(pattern, raw_text, re.MULTILINE)
            
            data_map = {}
            for match in matches:
                raw_name = match[0].strip().lower()
                clean_name = NAME_ALIASES.get(raw_name, raw_name)
                
                # यहाँ हम extract_float फंक्शन का यूज़ करके टेक्स्ट में से नंबर निकालेंगे
                data_map[clean_name] = {
                    'd': extract_float(match[1]), # जैसे "NIL" -> 0.0, " MT" -> 0.0
                    'm': extract_float(match[2]), # "1097.990 MT" -> 1097.990
                    'y': extract_float(match[3])
                }

            # ---------------------------------------------------------
            # PART D: Excel अपडेट
            # ---------------------------------------------------------
            updated_count = 0
            
            # Row 4 से शुरू
            for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_col=6), 4):
                name_cell = row[1]
                if name_cell.value:
                    excel_name = str(name_cell.value).strip().lower()
                    
                    # 1. पहले पुराना डेटा 0 करें (Reset)
                    if "description" not in excel_name and "date" not in excel_name:
                        ws.cell(row=row_idx, column=4).value = 0.0
                        ws.cell(row=row_idx, column=5).value = 0.0
                        ws.cell(row=row_idx, column=6).value = 0.0

                    # 2. नया डेटा भरें
                    if excel_name in data_map:
                        ws.cell(row=row_idx, column=4).value = data_map[excel_name]['d']
                        ws.cell(row=row_idx, column=5).value = data_map[excel_name]['m']
                        ws.cell(row=row_idx, column=6).value = data_map[excel_name]['y']
                        updated_count += 1

            # ---------------------------------------------------------
            # PART E: डाउनलोड
            # ---------------------------------------------------------
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            st.success(f"✅ अपडेटेड! {updated_count} एंट्रीज भरी गईं (NIL/Empty values handled).")
            st.download_button(
                label=f"📥 डाउनलोड DPR_{final_date_str}.xlsx",
                data=output,
                file_name=f"DPR_{final_date_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Error: {e}")
            
